"""Excel import / export utility based on openpyxl."""

from __future__ import annotations

import io
from collections.abc import Sequence
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Column descriptor
# ---------------------------------------------------------------------------
# Each element in *columns* is a dict with at least:
#   "header"  - 表头文字 (str)
#   "field"   - 数据字段名 (str)
# Optional:
#   "width"   - 列宽 (int, default auto)
#   "type"    - "str" | "int" | "float" | "date" (default "str")

_DEFAULT_COL_WIDTH = 18

_HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_header_style(ws, col_count: int) -> None:
    """Style the header row (row 1)."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _apply_data_style(ws, row_idx: int, col_count: int) -> None:
    """Style a single data row."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.alignment = _DATA_ALIGN
        cell.border = _THIN_BORDER


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def export_to_excel(
    data_list: Sequence[dict[str, Any]],
    columns: list[dict[str, Any]],
    filename: str | None = None,
) -> io.BytesIO:
    """Export *data_list* (list of dicts) to an Excel file.

    Parameters
    ----------
    data_list : list[dict]
        Each dict's keys should match ``columns[*]["field"]``.
    columns : list[dict]
        Column descriptors (see module docstring).
    filename : str | None
        If provided, the returned BytesIO will have ``.name`` set so
        FastAPI's ``StreamingResponse`` can infer the download name.

    Returns
    -------
    io.BytesIO
        In-memory Excel workbook (xlsx).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # -- Header row ----------------------------------------------------------
    for col_idx, col_def in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_def["header"])
        width = col_def.get("width", _DEFAULT_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    _apply_header_style(ws, len(columns))

    # -- Data rows -----------------------------------------------------------
    for row_offset, row_data in enumerate(data_list, start=2):
        for col_idx, col_def in enumerate(columns, start=1):
            value = row_data.get(col_def["field"], "")
            # Format dates for display
            col_type = col_def.get("type", "str")
            if col_type == "date" and isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=row_offset, column=col_idx, value=value)
        _apply_data_style(ws, row_offset, len(columns))

    # Freeze header
    ws.freeze_panes = "A2"

    # -- Write to buffer -----------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    if filename:
        buf.name = filename
    return buf


def import_from_excel(
    file_bytes: bytes,
    columns: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Parse an uploaded Excel file and return a list of dicts.

    The first row of the sheet must be the header.  Headers are matched
    against ``columns[*]["header"]`` (case-insensitive, stripped).

    Parameters
    ----------
    file_bytes : bytes
        Raw xlsx file content.
    columns : list[dict]
        Column descriptors.

    Returns
    -------
    list[dict]
        Parsed rows keyed by ``columns[*]["field"]``.
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
    ws = wb.active

    # Build header -> field mapping (case-insensitive)
    header_to_field: dict[str, str] = {col_def["header"].strip().lower(): col_def["field"] for col_def in columns}

    rows: list[dict[str, Any]] = []
    headers: list[str | None] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            # Header row: map each cell to its field name
            headers = [header_to_field.get(str(cell).strip().lower()) if cell else None for cell in row]
            continue

        row_dict: dict[str, Any] = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                field = headers[col_idx]
                assert field is not None  # mypy: narrowed by truthiness check above
                # Type coercion
                col_def = next((c for c in columns if c["field"] == field), None)
                col_type = col_def.get("type", "str") if col_def else "str"
                row_dict[field] = _coerce(cell, col_type)
        rows.append(row_dict)

    wb.close()
    return rows


def generate_template(columns: list[dict[str, Any]], filename: str | None = None) -> io.BytesIO:
    """Generate a blank import template with headers only.

    Returns
    -------
    io.BytesIO
        In-memory xlsx with styled header row and correct column widths.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"

    for col_idx, col_def in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_def["header"])
        width = col_def.get("width", _DEFAULT_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    _apply_header_style(ws, len(columns))

    # Add a sample hint row
    for col_idx, col_def in enumerate(columns, start=1):
        hint = _type_hint(col_def.get("type", "str"))
        cell = ws.cell(row=2, column=col_idx, value=hint)
        cell.font = Font(color="999999", italic=True)
        cell.alignment = _DATA_ALIGN

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    if filename:
        buf.name = filename
    return buf


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _coerce(value: Any, col_type: str) -> Any:
    """Coerce a cell value to the expected Python type."""
    if value is None:
        return None
    try:
        if col_type == "int":
            return int(value)
        if col_type == "float":
            return float(value)
        if col_type == "date":
            if isinstance(value, datetime):
                return value
            return datetime.strptime(str(value).strip(), "%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        pass
    return str(value).strip() if value is not None else ""


def _type_hint(col_type: str) -> str:
    """Return a human-readable hint for the template's sample row."""
    return {
        "str": "文本",
        "int": "整数",
        "float": "小数",
        "date": "日期 (YYYY-MM-DD HH:MM:SS)",
    }.get(col_type, "文本")
